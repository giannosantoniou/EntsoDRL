"""HEnEx IDA Clearing Price Collector.

I collect IDA1/IDA2/IDA3 clearing prices (Market Clearing Price, 15-min resolution)
from the Hellenic Energy Exchange (HEnEx / EnExGroup).

HEnEx publishes daily XLSX files with results for each IDA session:
  IDA1: gate closure D-1 15:00, covers all 96 quarter-hours of delivery day
  IDA2: gate closure D-1 22:00, covers all 96 quarter-hours of delivery day
  IDA3: gate closure D+0 10:00, covers quarter-hours 49-96 (13:00-24:00)

File structure (verified from 20260401_EL-IDA1_ResultsSummary_EN_v01.xlsx):
  Sheet 'MKT_Coupling' (or 'SPOT_Summary (SELL)'):
    Row 8: "Greece Mainland  (15min MCP)" — 96 clearing prices in columns B-CZ
    Row 2: Hour groups (1-24)
    Row 3: Quarter-hour numbers (1-96)
"""

import io
import logging
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import requests

from PredictFuturePrices.config import PATHS

logger = logging.getLogger(__name__)

# I define download URL templates for each IDA session
IDA_URL_TEMPLATES = {
    1: "https://www.enexgroup.gr/documents/20126/3257255/{date}_EL-IDA1_ResultsSummary_EN_v01.xlsx",
    2: "https://www.enexgroup.gr/documents/20126/3257274/{date}_EL-IDA2_ResultsSummary_EN_v01.xlsx",
    3: "https://www.enexgroup.gr/documents/20126/3257528/{date}_EL-IDA3_ResultsSummary_EN_v01.xlsx",
}

# I define the row label containing clearing prices
MCP_ROW_LABEL = "Greece Mainland"
MCP_ROW_QUALIFIER = "15min MCP"

# I define request settings
REQUEST_TIMEOUT = 30
REQUEST_DELAY = 1.0  # seconds between downloads


class HenexIDACollector:
    """I collect IDA1/2/3 clearing prices from HEnEx daily XLSX publications."""

    def __init__(self, output_path: Optional[Path] = None):
        self.name = "henex_ida_clearing"
        self.output_path = output_path or PATHS.ida_clearing_csv
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def get_default_start(self) -> datetime:
        """I return the start date for IDA data (IDAs launched 13 June 2024)."""
        return datetime(2024, 6, 13)

    def fetch_range(self, start: datetime, end: datetime) -> pd.DataFrame:
        """I fetch IDA clearing prices for a date range.

        For each day in the range, I download IDA1/2/3 XLSX files,
        extract the 15-min Market Clearing Price, and combine into
        a single DataFrame.

        Returns:
            DataFrame with DatetimeIndex (15-min), columns:
            ida1_clearing_price, ida2_clearing_price, ida3_clearing_price
        """
        all_frames = []
        current = start.date() if isinstance(start, datetime) else start
        end_date = end.date() if isinstance(end, datetime) else end

        while current <= end_date:
            date_str = current.strftime("%Y%m%d")
            day_data = self._fetch_day(current, date_str)

            if day_data is not None and not day_data.empty:
                all_frames.append(day_data)

            current += timedelta(days=1)
            time.sleep(REQUEST_DELAY)

        if not all_frames:
            logger.warning(f"[{self.name}] No data fetched for {start} - {end}")
            return pd.DataFrame()

        result = pd.concat(all_frames).sort_index()
        result = result[~result.index.duplicated(keep='last')]
        result.index.name = "timestamp"
        return result

    def _fetch_day(self, date, date_str: str) -> Optional[pd.DataFrame]:
        """I fetch all 3 IDA sessions for one day and combine."""
        day_frames = {}

        for ida_num in [1, 2, 3]:
            url = IDA_URL_TEMPLATES[ida_num].format(date=date_str)
            try:
                resp = requests.get(url, timeout=REQUEST_TIMEOUT)
                if resp.status_code == 200:
                    prices = self._parse_ida_xlsx(resp.content, ida_num, date)
                    if prices is not None:
                        day_frames[f'ida{ida_num}_clearing_price'] = prices
                        logger.debug(f"[{self.name}] IDA{ida_num} {date_str}: {len(prices)} prices")
                elif resp.status_code == 404:
                    logger.debug(f"[{self.name}] IDA{ida_num} {date_str}: not published yet")
                else:
                    logger.warning(f"[{self.name}] IDA{ida_num} {date_str}: HTTP {resp.status_code}")
            except requests.RequestException as e:
                logger.warning(f"[{self.name}] IDA{ida_num} {date_str}: {e}")

            time.sleep(0.3)  # I wait between IDA sessions

        if not day_frames:
            return None

        # I build 15-min timestamps for this day
        timestamps = pd.date_range(
            start=pd.Timestamp(date),
            periods=96,
            freq='15min'
        )

        df = pd.DataFrame(index=timestamps)
        for col, prices in day_frames.items():
            if len(prices) >= 96:
                df[col] = prices[:96]
            else:
                # I pad with NaN for incomplete data (IDA3 only covers afternoon)
                padded = np.full(96, np.nan)
                # IDA3 starts at slot 49 (hour 13), but file may have different structure
                if 'ida3' in col and len(prices) < 96:
                    padded[96 - len(prices):] = prices
                else:
                    padded[:len(prices)] = prices
                df[col] = padded

        return df

    def _parse_ida_xlsx(self, content: bytes, ida_num: int, date) -> Optional[np.ndarray]:
        """I extract 15-min MCP from an IDA ResultsSummary XLSX file.

        I search for the row containing "Greece Mainland" and "15min MCP"
        and extract the clearing prices from columns B onwards (96 values).
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                io.BytesIO(content), data_only=True, read_only=True
            )

            # I try multiple sheet names (structure may vary)
            sheet_names_to_try = ['MKT_Coupling', 'SPOT_Summary (SELL)', 'SPOT_Summary (BUY)']
            ws = None
            for sheet_name in sheet_names_to_try:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    break

            if ws is None:
                # I fall back to first sheet
                ws = wb[wb.sheetnames[0]]

            # I search for the MCP row (row 8 in known files, but I search by label)
            mcp_row = None
            for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=1, values_only=False):
                cell = row[0]
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if MCP_ROW_LABEL in val and MCP_ROW_QUALIFIER in val:
                        mcp_row = cell.row
                        break

            if mcp_row is None:
                logger.warning(f"[{self.name}] IDA{ida_num} {date}: MCP row not found")
                wb.close()
                return None

            # I extract prices from columns B onwards (col 2 to 97 = 96 values)
            prices = []
            for col in range(2, 98):  # B=2 to CU=97
                cell = ws.cell(row=mcp_row, column=col)
                if cell.value is not None:
                    try:
                        prices.append(float(cell.value))
                    except (ValueError, TypeError):
                        prices.append(np.nan)
                else:
                    prices.append(np.nan)

            wb.close()

            if len(prices) == 0:
                return None

            return np.array(prices)

        except Exception as e:
            logger.error(f"[{self.name}] IDA{ida_num} {date}: Parse error: {e}")
            return None

    def fetch_incremental(self, end: Optional[datetime] = None) -> pd.DataFrame:
        """I fetch only new data since last collection."""
        if end is None:
            end = datetime.now()

        last = self._get_last_timestamp()
        if last is not None:
            start = last + timedelta(hours=1)
        else:
            start = self.get_default_start()

        if start >= end:
            logger.info(f"[{self.name}] Already up to date.")
            return pd.DataFrame()

        logger.info(f"[{self.name}] Fetching {start.date()} -> {end.date()}")

        try:
            new_data = self.fetch_range(start, end)
        except Exception as e:
            logger.error(f"[{self.name}] Fetch failed: {e}")
            return pd.DataFrame()

        if not new_data.empty:
            self._save_incremental(new_data)

        return new_data

    def _get_last_timestamp(self) -> Optional[datetime]:
        """I read the last timestamp from the CSV file."""
        if not self.output_path.exists():
            return None
        try:
            df = pd.read_csv(self.output_path, index_col=0, parse_dates=True)
            if df.empty:
                return None
            last = pd.Timestamp(df.index.max())
            if last.tzinfo is not None:
                last = last.tz_localize(None)
            return last.to_pydatetime()
        except Exception:
            return None

    def _save_incremental(self, new_data: pd.DataFrame) -> None:
        """I merge new data with existing CSV and save."""
        new_data.index.name = "timestamp"
        if new_data.index.tz is not None:
            new_data.index = new_data.index.tz_localize(None)

        if self.output_path.exists():
            existing = pd.read_csv(self.output_path, index_col=0, parse_dates=True)
            if existing.index.tz is not None:
                existing.index = existing.index.tz_localize(None)
            merged = pd.concat([existing, new_data])
            merged = merged[~merged.index.duplicated(keep="last")]
            merged = merged.sort_index()
        else:
            merged = new_data.sort_index()

        merged.to_csv(self.output_path)
        logger.info(f"[{self.name}] Saved {len(merged)} rows to {self.output_path.name}")
